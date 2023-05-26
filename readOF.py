import datetime
import logging
import os
import pandas as pd
import pythoncom
import win32com.client as win32
import openpyxl
from openpyxl.utils import get_column_letter
import config
import pickle
import time


def get_project(path):
    """Открывает файл проекта и возвращает объект проекта"""
    if not os.path.isabs(path):
        logging.warning('%s: Путь до файла проекта не абсолютный', get_project.__name__)
    logging.info('%s: Пытаемся открыть файл проекта', get_project.__name__)
    try:
        msp = win32.Dispatch("MSProject.Application", pythoncom.CoInitialize())
        _abs_path = os.path.abspath(path)
        print(_abs_path)
        msp.FileOpen(_abs_path)
        project = msp.ActiveProject
    except Exception:
        logging.error('%s: Файл проекта не смог открыться', get_project.__name__)
        raise Exception('Не получилось открыть файл проекта')
    logging.info('%s: Файл проекта успешно открылся', get_project.__name__)
    return project, msp


def get_data_task(t, msp):
    """"Получает значения task из нужных столбцов"""
    arr = []
    try:
        for i in config.id_column.keys():
            data = getattr(t,i)
            if isinstance(data, datetime.datetime):
                data = datetime.datetime.date(data)
            arr.append(data)
    except Exception as e:
        print(e)
        logging.error('%s: Неверный идентификатор столбца project', get_project.__name__)
        raise Exception('Неверный идентификатор столбца project')
    return arr


def fill_dataframe(project, msp):
    """Заполняет DataFrame значениями из project"""
    logging.info('%s: Создаем DataFrame из столбцов объекта проекта', fill_dataframe.__name__)
    if not project:
        logging.error('%s: Не удалось получить объект проекта', fill_dataframe.__name__)
        raise Exception("Объект проекта пустой")
    if not config.id_column:
        logging.error('%s: Ключевые столбцы не заданы', fill_dataframe.__name__)
        raise Exception("Ключевые столбцы не заданы")
    task_collection = project.Tasks
    data = pd.DataFrame(columns=config.id_column.values())
    start = time.time()
    try:
        for t in task_collection:
            data.loc[len(data.index)] = get_data_task(t, msp)
    except Exception:
        logging.error('%s: Неверно заполнен словарь столбцов и их идентификаторов', fill_dataframe.__name__)
        raise Exception("Ошибка в словаре слобцов и их идентификаторов")
    logging.info('%s: DataFrame из столбцов объекта проекта успешно создан', fill_dataframe.__name__)

    return data


def set_column_date_format(worksheet, column_index, date_format):
    """Пытается присвоить формат Дата столбцу в Excel"""
    column_letter = get_column_letter(column_index)
    for cell in worksheet[column_letter]:
        cell.number_format = date_format



def set_style_excel(column_index, path_to_excel, columns_to_convert, data):
    """Применяет стили к строкам excel"""
    try:
        with open(config.path_to_style_file, 'rb') as file:
            styles_dict = pickle.load(file)
    except FileNotFoundError:
        logging.error('%s: Неверно задан путь к файлу со стилями', fill_dataframe.__name__)
        raise Exception("Неверный путь до файла со стилями")
    workbook_other = openpyxl.load_workbook(path_to_excel)
    worksheet_other = workbook_other.active
    for row in worksheet_other.iter_rows(min_row=0):
        cell = row[column_index - 1]
        cell_value = cell.value
        if cell_value in styles_dict:
            for i in range(0, worksheet_other.max_column):
                cell = row[i]
                style = styles_dict[cell_value]
                cell.style = style
                column_letter = get_column_letter(cell.column)
                text_length = len(str(cell.value))
                current_width = worksheet_other.column_dimensions[column_letter].width
                if text_length > current_width:
                    worksheet_other.column_dimensions[column_letter].width = text_length
    date_format = 'dd.mm.yyyy'
    for column in columns_to_convert:
        set_column_date_format(worksheet_other, data.columns.get_loc(column) + 1, date_format)
    workbook_other.save(path_to_excel)


def main(path_to_project, path_to_folder):
    """Управляющая функция"""
    path_to_excel = None
    try:
        start = time.time()
        project, msp = get_project(path_to_project)
        data = fill_dataframe(project, msp)
        file_name = os.path.splitext(os.path.basename(path_to_project))[0]
        current_date = datetime.datetime.now().strftime("%d.%m.%Y")
        path_to_excel = path_to_folder + "//" + file_name + "_ОФ_" + current_date + ".xlsx"
        columns_to_convert = data.columns[data.columns.str.contains('начало|окончание', case=False)]
        data.to_excel(path_to_excel, index=False)
        column_index = data.columns.get_loc(config.id_column['Text5']) + 1
        set_style_excel(column_index, path_to_excel, columns_to_convert, data)
        end = time.time()
        print(end - start)
    except Exception as e:
        print(e)
        return path_to_excel
    return path_to_excel

